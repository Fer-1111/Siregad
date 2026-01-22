# extractor.py

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
import pandas as pd
import unicodedata


def extraer_movimientos(path_excel, config, fecha, division):
    registros = []

    if hasattr(path_excel, "seek"):
        path_excel.seek(0)

    wb = load_workbook(path_excel, data_only=True)

    # Debug: mostrar hojas disponibles
    hojas_disponibles = wb.sheetnames
    print(f"[DEBUG] División: {division}")
    print(f"[DEBUG] Hojas disponibles: {hojas_disponibles}")

    def parse_numeric(v):
        """Try to convert v to float. Return None if not numeric."""
        if v is None:
            return None
        if isinstance(v, (int, float)):
            try:
                return float(v)
            except Exception:
                return None
        if isinstance(v, str):
            s = v.strip()
            if s == "":
                return None
            # normalize decimal comma
            s = s.replace(" ", "")
            s = s.replace(",", ".")
            try:
                return float(s)
            except Exception:
                return None
        return None

    def normalize_sheet_name(s: str) -> str:
        if s is None:
            return ""
        s2 = str(s).strip().lower()
        s2 = unicodedata.normalize('NFKD', s2)
        s2 = ''.join(c for c in s2 if not unicodedata.combining(c))
        s2 = s2.replace(' ', '')
        return s2

    def find_sheet_name(requested, available_names):
        if requested in available_names:
            return requested
        # try exact lower match
        for name in available_names:
            if name.lower() == requested.lower():
                return name
        # try normalized match (remove accents and spaces)
        req_norm = normalize_sheet_name(requested)
        for name in available_names:
            if normalize_sheet_name(name) == req_norm:
                return name
        # try contains
        for name in available_names:
            if req_norm in normalize_sheet_name(name) or normalize_sheet_name(name) in req_norm:
                return name
        return None

    for concepto, cfg in config.items():
        try:
            hoja_requerida = cfg["sheet"]

            # Verificar si la hoja existe (tolerante a mayúsculas/acentos/espacios)
            matched = find_sheet_name(hoja_requerida, hojas_disponibles)
            if not matched:
                print(f"[DEBUG] WARNING: La hoja solicitada '{hoja_requerida}' NO existe. Hojas: {hojas_disponibles}")
                # Saltar este concepto en lugar de fallar toda la extracción
                continue

            ws = wb[matched]

            # Detectar si es un rango (ej: K10:K13) o una celda simple
            celda = cfg["cell"]
            valor = None
            if ":" in celda:
                # Es un rango, sumar todas las celdas de forma eficiente usando iter_rows
                min_col, min_row, max_col, max_row = range_boundaries(celda)
                suma = 0.0
                found = False
                for row_vals in ws.iter_rows(min_row=min_row, max_row=max_row,
                                             min_col=min_col, max_col=max_col,
                                             values_only=True):
                    for cell_val in row_vals:
                        n = parse_numeric(cell_val)
                        if n is not None:
                            suma += n
                            found = True
                valor = suma if found else None
            else:
                # Es una celda simple, obtener por coordenadas y parsear
                col_letter, row = coordinate_from_string(celda)
                col = column_index_from_string(col_letter)
                cell_obj = ws.cell(row=row, column=col)
                cell_val = cell_obj.value
                # Debug detallado para diagnóstico con tipo de dato y fórmula si existe
                cell_type = type(cell_val).__name__
                has_formula = hasattr(cell_obj, 'data_type') and cell_obj.data_type == 'f'
                debug_msg = f"📍 {concepto}: Hoja='{matched}' | Celda={celda} (Row{row}Col{col}) | Tipo={cell_type} | Raw={repr(cell_val)}"
                if has_formula:
                    debug_msg += f" | Fórmula detectada"
                print(debug_msg)
                valor = parse_numeric(cell_val) if cell_val is not None else None
                print(f"   → Parseado: {valor}")

            if valor is not None:
                print(f"✓ {concepto}: {celda} = {valor}")

            # Saltar si el valor es None o 0 (pero incluir si es inventario
            # o si la configuración indica explícitamente incluir ceros)
            if valor is None or valor == 0:
                if "inventario" in concepto or cfg.get("include_if_zero", False):
                    valor = 0  # Forzar 0 cuando se requiere incluir
                else:
                    # No hay un valor numérico en la celda; saltar
                    continue

            # Guardar todo en registro intermedio
            registro = {
                "Division": division,
                "Concepto": concepto,
                "Grupo": cfg.get("grupo", concepto),  # Usar grupo si existe, sino usar concepto
                "Fecha": fecha,
                "Cantidad": abs(float(valor)),  # Valor absoluto
                "Unidad": "TN",
                "IncludeBatch": cfg.get("include_in_batch", True)
            }

            # Solo agregar campos de batch si va incluido
            if cfg.get("include_in_batch", True):
                registro.update({
                    "Bodega": cfg.get("bodega", ""),
                    "Material": cfg.get("material", ""),
                    "Tipo Movimiento": cfg.get("tipo_mov", ""),
                    "Movimiento": cfg.get("movimiento", ""),
                })
            else:
                # Para registros que no van al batch, poner valores vacíos
                registro.update({
                    "Bodega": "",
                    "Material": "",
                    "Tipo Movimiento": "",
                    "Movimiento": "",
                })

            registros.append(registro)
        except Exception as e:
            raise Exception(f"Error en concepto '{concepto}' (hoja: {cfg.get('sheet')}, celda: {cfg.get('cell')}): {str(e)}")

    return pd.DataFrame(registros)


def extraer_siregad(path_excel):
    """
    Extrae datos de la hoja 'Batch completo' del archivo SIREGAD
    """
    if hasattr(path_excel, "seek"):
        path_excel.seek(0)
    
    try:
        # Leer la hoja "Batch completo"
        df = pd.read_excel(path_excel, sheet_name="Batch completo")
        
        # Columnas relevantes (ajustar según estructura real)
        columnas_relevantes = [
            "Número Factura", "Id. Bodega", "SQC CAS/Nombre Mezcla", 
            "Concentración (0,00000)", "Ingreso/Egreso", "Tipo de Movimiento",
            "Cantidad", "Unidad Medida", "Expediente Comunicación Anticipada",
            "Rut", "Nombre"
        ]
        
        # Filtrar solo las columnas que existen
        columnas_disponibles = [col for col in columnas_relevantes if col in df.columns]
        
        if columnas_disponibles:
            df_filtrado = df[columnas_disponibles].copy()
        else:
            df_filtrado = df.copy()
        
        # Eliminar filas completamente vacías
        df_filtrado = df_filtrado.dropna(how='all')
        
        return df_filtrado
    
    except Exception as e:
        raise Exception(f"Error al extraer SIREGAD: {str(e)}")
