import io
from datetime import date

import streamlit as st
import pandas as pd

from batch_writer import escribir_batch
# Barquito es opcional: si el archivo de config está desactualizado, evitamos que la app falle
try:
    from config import CONFIG_POTRERILLOS, CONFIG_CALETONES, CONFIG_TBA, CONFIG_SAN_ANTONIO, CONFIG_RT, CONFIG_CHUQUICAMATA, CONFIG_DMH, CONFIG_BARQUITO, CONFIG_DGM
except ImportError:
    from config import CONFIG_POTRERILLOS, CONFIG_CALETONES, CONFIG_TBA, CONFIG_SAN_ANTONIO, CONFIG_RT, CONFIG_CHUQUICAMATA, CONFIG_DMH
    CONFIG_BARQUITO = {}
    CONFIG_DGM = {}
from config_programacion import (
    COLUMNA_MES_D_START, MES_ALIAS, DIVISION_A_GRUPO_PROGRA, TOLERANCIA_DIFERENCIA,
    CONFIG_SALV, CONFIG_TTE, CONFIG_DMH_PROGRA, CONFIG_CHUQUI, CONFIG_VENT,
    CONFIG_MEJ, CONFIG_MEJICL, CONFIG_TERMEJ, CONFIG_COMPRAS
)
from extractor import extraer_movimientos, extraer_siregad
from validator import validar_inventario

st.set_page_config(
    page_title="Carga SIREGAD",
    page_icon="📦",
    layout="wide",
)

st.title("📦 Generador Batch SIREGAD")

# Mapeo de palabras clave en nombres de archivo a configuraciones
DIVISIONES_CONFIG = {
    "Potrerillos": CONFIG_POTRERILLOS,
    "Caletones": CONFIG_CALETONES,
    "TBA": CONFIG_TBA,
    "San Antonio": CONFIG_SAN_ANTONIO,
    "Radomiro Tomic": CONFIG_RT,
    "Chuquicamata": CONFIG_CHUQUICAMATA,
    "Ministro Hales": CONFIG_DMH,
    "Barquito": CONFIG_BARQUITO,
    "Gabriela Mistral": CONFIG_DGM,
}

# Mapeo de palabras clave para detectar división en nombre de archivo
# IMPORTANTE: El orden importa - las más específicas deben ir primero
KEYWORDS_DIVISION = {
    "san antonio": "San Antonio",    # Más específico
    "sanantonio": "San Antonio",
    "rt -": "Radomiro Tomic",        # Más específico que solo "rt"
    "dgm": "Gabriela Mistral",       # DGM = Gabriela Mistral
    "dmh": "Ministro Hales",         # Antes de "dch" para evitar conflictos
    "dch": "Chuquicamata",
    "salvador": "Salvador",
    "potrerillos": "Potrerillos",
    "potre": "Potrerillos",
    "caletones": "Caletones",
    "calet": "Caletones",
    "calpo": "Caletones",
    "teca": "Caletones",
    "tba": "TBA",
    "tbape": "TBA",
    "chuquicamata": "Chuquicamata",
    "chuqui": "Chuquicamata",
    "radomiro": "Radomiro Tomic",
    "tomic": "Radomiro Tomic",
    "rt": "Radomiro Tomic",
    "gabriela": "Gabriela Mistral",
    "mistral": "Gabriela Mistral",
    "ministro": "Ministro Hales",
    "hales": "Ministro Hales",
    "ventana": "Ventana",
    "barquito": "Barquito",
    "tas": "Barquito",
}

# Nota: Caletones, TBA y San Antonio usan el mismo Excel (hoja "Balance") con columnas diferentes


def detectar_division(nombre_archivo):
    """Detecta la división basándose en el nombre del archivo"""
    nombre_lower = nombre_archivo.lower()
    for keyword, division in KEYWORDS_DIVISION.items():
        if keyword in nombre_lower:
            return division
    return None


# Calcular el último día del mes anterior al mes actual
import calendar
hoy = date.today()

# Si estamos en enero, el mes anterior es diciembre del año anterior
if hoy.month == 1:
    mes_seleccionado = 12
    anio_seleccionado = hoy.year - 1
else:
    mes_seleccionado = hoy.month - 1
    anio_seleccionado = hoy.year

ultimo_dia = calendar.monthrange(anio_seleccionado, mes_seleccionado)[1]
fecha_ultimo_dia = date(anio_seleccionado, mes_seleccionado, ultimo_dia)
fecha = fecha_ultimo_dia

MESES_NOMBRES = {1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril", 5: "Mayo", 6: "Junio",
                 7: "Julio", 8: "Agosto", 9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"}
st.info(f"📅 Procesando: **{MESES_NOMBRES[mes_seleccionado]} {anio_seleccionado}** (Fecha: {fecha_ultimo_dia.strftime('%d-%m-%Y')})")

# ============================================================================
# FASE 0: CARGAR ARCHIVO EXPORT (SOC) - FUENTE MAESTRA DE FACTURAS
# ============================================================================
st.subheader("📊 Fase 0: Cargar Archivo Export (SOC)")

with st.expander("📤 Cargar facturas desde Export SOC", expanded=True):
    st.markdown("""
    **Este es el archivo maestro de facturas exportado desde SAP/SOC.**
    
    📌 **Columnas esperadas:**
    - **Contrato Nuevo** (A): Código como CL5K402, ZO5K404, ZR5K311...
    - **Fecha Real SM** (Q): Para filtrar por mes
    - **Peso Neto**: Cantidad en TM
    - **Almacén**: Origen/destino (Plta.Los Lirios, Bod.Ventanas, etc.)
    - **Gl.Cliente**: Nombre del cliente
    
    📌 **Clasificación automática por prefijo:**
    - **CL** = Venta Nacional → código VENT
    - **ZO** = Canje salida (egreso)
    - **ZR** = Canje entrada (ingreso)
    """)
    
    archivo_export = st.file_uploader(
        "Sube el archivo Export (SOC)",
        type=["xlsx", "xls"],
        key="export_soc"
    )
    
    if archivo_export:
        try:
            # Leer el archivo
            df_export = pd.read_excel(archivo_export)
            
            st.write(f"📋 Columnas encontradas: {list(df_export.columns)}")
            st.write(f"📊 Total registros: {len(df_export)}")
            
            # Detectar columna de fecha (buscar "Fecha Real SM" o columna Q)
            col_fecha = None
            for col in df_export.columns:
                if "fecha" in col.lower() and "real" in col.lower():
                    col_fecha = col
                    break
            
            if col_fecha is None and len(df_export.columns) >= 17:
                col_fecha = df_export.columns[16]  # Columna Q (índice 16)
                st.info(f"Usando columna '{col_fecha}' como fecha")
            
            # Detectar columna de contrato
            col_contrato = None
            for col in df_export.columns:
                if "contrato" in col.lower():
                    col_contrato = col
                    break
            if col_contrato is None:
                col_contrato = df_export.columns[0]  # Primera columna
            
            # Detectar columna de cantidad (Peso Neto)
            col_cantidad = None
            for col in df_export.columns:
                if "peso" in col.lower() and "neto" in col.lower():
                    col_cantidad = col
                    break
            
            # Detectar columna de almacén
            col_almacen = None
            for col in df_export.columns:
                if "almac" in col.lower():
                    col_almacen = col
                    break
            
            # Detectar columna de cliente
            col_cliente = None
            for col in df_export.columns:
                if "cliente" in col.lower() or "gl.cliente" in col.lower():
                    col_cliente = col
                    break
            
            st.write(f"🔍 Columnas detectadas: Contrato='{col_contrato}', Fecha='{col_fecha}', Cantidad='{col_cantidad}', Almacén='{col_almacen}'")
            
            # Filtrar por mes si hay columna de fecha
            if col_fecha and st.checkbox("Filtrar por mes seleccionado", value=True, key="filtrar_export"):
                df_export[col_fecha] = pd.to_datetime(df_export[col_fecha], errors='coerce')
                df_filtrado = df_export[
                    (df_export[col_fecha].dt.month == mes_seleccionado) &
                    (df_export[col_fecha].dt.year == anio_seleccionado)
                ]
                st.success(f"✅ Registros del mes: {len(df_filtrado)} de {len(df_export)}")
            else:
                df_filtrado = df_export
            
            if len(df_filtrado) > 0 and st.button("📊 Procesar y Clasificar", key="btn_procesar_export"):
                # Clasificar por tipo de contrato
                def clasificar_contrato(contrato):
                    if pd.isna(contrato):
                        return "OTRO", "?"
                    contrato = str(contrato).upper()
                    if contrato.startswith("CL"):
                        return "VENT", "E"  # Venta nacional - Egreso
                    elif contrato.startswith("ZO"):
                        return "EDEV", "E"  # Canje salida - Egreso (devolución)
                    elif contrato.startswith("ZR"):
                        return "MDEV", "I"  # Canje entrada - Ingreso (recepción)
                    elif contrato.startswith("ZU"):
                        return "IM", "I"    # Importación
                    else:
                        return "OTRO", "?"
                
                # Mapeo de Almacén a División/Bodega
                ALMACEN_A_DIVISION = {
                    "plta.los lirios": ("TBA", "TETBA"),
                    "los lirios": ("TBA", "TETBA"),
                    "bod.ventanas": ("Ventana", "VENT"),
                    "ventanas": ("Ventana", "VENT"),
                    "tas barquito": ("Barquito", "BARQUITO"),
                    "barquito": ("Barquito", "BARQUITO"),
                    "term.mejillones": ("Terquim", "TERMEJ"),
                    "terquim mejillones": ("Terquim", "TERMEJ"),
                    "terquim mejillon": ("Terquim", "TERMEJ"),
                    "interacid mej": ("Mejillones ICL", "MEJICL"),
                    "interacid mej.": ("Mejillones ICL", "MEJICL"),
                    "mejillones icl": ("Mejillones ICL", "MEJICL"),
                    "bodega trading": ("Potrerillos", "SALPO"),
                    "potrerillos": ("Potrerillos", "SALPO"),
                    "caletones": ("Caletones", "TECA"),
                    "san antonio": ("San Antonio", "CM"),
                }
                
                def obtener_division_bodega(almacen):
                    if pd.isna(almacen):
                        return "Desconocido", "?"
                    almacen_lower = str(almacen).lower().strip()
                    for key, (div, bod) in ALMACEN_A_DIVISION.items():
                        if key in almacen_lower:
                            return div, bod
                    return str(almacen), "?"
                
                # Procesar cada registro
                registros_procesados = []
                for idx, row in df_filtrado.iterrows():
                    contrato = row[col_contrato] if col_contrato else ""
                    tipo_mov, ingreso_egreso = clasificar_contrato(contrato)
                    
                    almacen = row[col_almacen] if col_almacen else ""
                    division, bodega = obtener_division_bodega(almacen)
                    
                    cantidad = row[col_cantidad] if col_cantidad else 0
                    try:
                        cantidad = float(str(cantidad).replace(",", ".").replace(" ", "")) if cantidad else 0
                    except:
                        cantidad = 0
                    
                    cliente = row[col_cliente] if col_cliente else ""
                    
                    if cantidad > 0:
                        registros_procesados.append({
                            "Contrato": contrato,
                            "Cliente": cliente,
                            "Cantidad": cantidad,
                            "Almacén": almacen,
                            "División": division,
                            "Bodega": bodega,
                            "Tipo Mov": tipo_mov,
                            "I/E": ingreso_egreso,
                        })
                
                df_procesado = pd.DataFrame(registros_procesados)
                
                # Mostrar resumen
                st.success(f"✅ Procesados {len(df_procesado)} registros")
                
                # Resumen por tipo
                col1, col2, col3 = st.columns(3)
                with col1:
                    ventas = df_procesado[df_procesado["Tipo Mov"] == "VENT"]["Cantidad"].sum()
                    st.metric("💰 Ventas (VENT)", f"{ventas:,.2f} TM")
                with col2:
                    canjes_sal = df_procesado[df_procesado["Tipo Mov"] == "EDEV"]["Cantidad"].sum()
                    st.metric("↗️ Canjes Salida (ZO)", f"{canjes_sal:,.2f} TM")
                with col3:
                    canjes_ent = df_procesado[df_procesado["Tipo Mov"] == "MDEV"]["Cantidad"].sum()
                    st.metric("↙️ Canjes Entrada (ZR)", f"{canjes_ent:,.2f} TM")
                
                # Resumen por división
                st.write("### 📊 Resumen por División")
                resumen_div = df_procesado.groupby(["División", "Tipo Mov"]).agg({
                    "Cantidad": "sum",
                    "Contrato": "count"
                }).reset_index()
                resumen_div.columns = ["División", "Tipo Mov", "Total TM", "Nº Registros"]
                st.dataframe(resumen_div, use_container_width=True)
                
                # Mostrar detalle
                with st.expander("📋 Ver detalle de registros", expanded=False):
                    st.dataframe(df_procesado, use_container_width=True)
                
                # Guardar en session_state
                st.session_state.df_export_procesado = df_procesado
                st.session_state.export_cargado = True
                
        except Exception as e:
            st.error(f"Error al leer archivo: {e}")
            import traceback
            st.code(traceback.format_exc())

# Mostrar datos de Export si ya están cargados
if "df_export_procesado" in st.session_state:
    with st.expander("📊 Datos Export cargados", expanded=False):
        df_exp = st.session_state.df_export_procesado
        st.write(f"Total registros: {len(df_exp)}")
        resumen = df_exp.groupby("Tipo Mov")["Cantidad"].sum()
        st.dataframe(resumen, use_container_width=True)

st.markdown("---")

# ============================================================================
# FASE 1: CARGAR BALANCES DE DIVISIONES
# ============================================================================
st.subheader("📂 Fase 1: Cargar Balances de Divisiones")

# Mostrar convenciones de nombres
with st.expander("ℹ️ Convención de nombres de archivos", expanded=False):
    st.markdown("""
    **El nombre del archivo debe contener una de estas palabras clave:**
    
    - `Salvador` o `SALPO` → División Salvador
    - `Potrerillos` o `Potre` → División Potrerillos  
    - `Caletones` o `CALPO` o `TECA` → División Caletones
    - `TBA` o `TBAPE` → División TBA
    - `Chuquicamata` o `Chuqui` → División Chuquicamata
    - `Radomiro` o `Tomic` → División Radomiro Tomic
    - `Gabriela` o `Mistral` → División Gabriela Mistral
    - `Ministro` o `Hales` → División Ministro Hales
    - `Ventana` → División Ventana
    - `Barquito` o `TAS` → División Barquito
    - `DGM` → División Gabriela Mistral
    """)

# Cargar múltiples archivos a la vez
archivos = st.file_uploader(
    "Sube todos los Excel de las divisiones",
    type=["xlsx"],
    accept_multiple_files=True,
    key="balances"
)

st.subheader("📋 Fase 1.5: Cargar Respaldo SIREGAD (Opcional)")

archivo_siregad = st.file_uploader(
    "Sube el archivo SIREGAD Noviembre para auditoría",
    type=["xlsx"],
    key="siregad"
)

# ===== PASO 2.5: IMPORTAR FACTURAS DESDE EXCEL CM =====
st.subheader("📑 Fase 2: Importar Facturas CM (Ventas/Compras)")

with st.expander("📤 Importar desde Excel Casa Matriz (CM)", expanded=False):
    st.markdown("""
    **Este importador lee el Excel de CM con las hojas:**
    - **Ventas**: Egresos (VENT, CL) - columnas I-U con formato batch
    - **Compras**: Ingresos (CL, IM) - columnas I-U con formato batch
    
    Las columnas I-U deben tener: `Número Fact | Id. Bodega | SQC CAS | Concentración | Ingreso/Egreso | Tipo de Mov | Cantidad | Unidad Med | Expediente | Rut | Nombre`
    """)
    
    archivo_cm = st.file_uploader(
        "Sube Excel de Casa Matriz (con hojas Ventas y Compras)",
        type=["xlsx"],
        key="cm_import"
    )
    
    col1, col2 = st.columns(2)
    with col1:
        importar_ventas = st.checkbox("Importar Ventas (Egresos)", value=True)
    with col2:
        importar_compras = st.checkbox("Importar Compras (Ingresos)", value=True)
    
    # División para los datos importados
    division_cm = st.selectbox(
        "División para estos datos",
        ["San Antonio", "Ministro Hales"],
        key="div_cm",
        help="CM = Casa Matriz, puede ser San Antonio o Ministro Hales"
    )
    
    if archivo_cm and st.button("📥 Importar desde CM"):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(archivo_cm, data_only=True)
            hojas_disponibles = wb.sheetnames
            st.write(f"📋 Hojas encontradas: {hojas_disponibles}")
            
            registros_importados = []
            
            # Función para leer hoja con formato CM
            def leer_hoja_cm(wb, nombre_hoja, division):
                registros = []
                if nombre_hoja not in wb.sheetnames:
                    st.warning(f"⚠️ Hoja '{nombre_hoja}' no encontrada")
                    return registros
                
                ws = wb[nombre_hoja]
                
                # Las columnas I-U contienen los datos del batch (columnas 9-21)
                # I=9: Número Fact, J=10: Id. Bodega, K=11: SQC CAS, L=12: Concentración
                # M=13: Ingreso/Egreso, N=14: Tipo de Mov, O=15: Cantidad, P=16: Unidad Med
                # Q=17: Expediente, R=18: Rut, S=19: Nombre
                
                for row in range(2, ws.max_row + 1):  # Empezar desde fila 2 (saltando header)
                    num_factura = ws.cell(row=row, column=9).value  # Columna I
                    cantidad = ws.cell(row=row, column=15).value    # Columna O
                    
                    # Convertir cantidad a número
                    try:
                        cantidad_num = float(str(cantidad).replace(",", ".").replace(" ", "")) if cantidad else 0
                    except:
                        cantidad_num = 0
                    
                    # Solo importar si hay factura y cantidad > 0
                    if num_factura and cantidad_num > 0:
                        registro = {
                            "Division": division,
                            "Concepto": str(num_factura),
                            "Grupo": str(num_factura),
                            "Bodega": ws.cell(row=row, column=10).value or "CM",  # J
                            "Material": ws.cell(row=row, column=11).value or "7664-93-9",  # K
                            "Concentracion": ws.cell(row=row, column=12).value or "Ácido Sulfúrico (Conc: 96)",  # L
                            "Movimiento": ws.cell(row=row, column=13).value or "",  # M (I/E)
                            "Tipo Movimiento": ws.cell(row=row, column=14).value or "",  # N
                            "Cantidad": cantidad_num,
                            "Unidad": ws.cell(row=row, column=16).value or "TN",  # P
                            "Expediente": ws.cell(row=row, column=17).value or "",  # Q
                            "Rut": ws.cell(row=row, column=18).value or "",  # R
                            "Nombre": ws.cell(row=row, column=19).value or "",  # S
                            "IncludeBatch": True,
                            "Fecha": fecha,
                        }
                        registros.append(registro)
                
                return registros
            
            # Importar Ventas (Egresos)
            if importar_ventas:
                ventas = leer_hoja_cm(wb, "Ventas", division_cm)
                if ventas:
                    registros_importados.extend(ventas)
                    st.success(f"✅ Ventas: {len(ventas)} registros importados")
            
            # Importar Compras (Ingresos)
            if importar_compras:
                compras = leer_hoja_cm(wb, "Compras", division_cm)
                if compras:
                    registros_importados.extend(compras)
                    st.success(f"✅ Compras: {len(compras)} registros importados")
            
            wb.close()
            
            if registros_importados:
                # Guardar en session_state
                if "facturas_importadas" not in st.session_state:
                    st.session_state.facturas_importadas = []
                
                st.session_state.facturas_importadas.extend(registros_importados)
                
                # Mostrar resumen
                df_preview = pd.DataFrame(registros_importados)
                st.write("**📊 Resumen de importación:**")
                resumen = df_preview.groupby(["Tipo Movimiento", "Movimiento"]).agg({
                    "Cantidad": ["count", "sum"]
                }).reset_index()
                resumen.columns = ["Tipo Mov", "I/E", "Registros", "Total TN"]
                st.dataframe(resumen, use_container_width=True)
                
                st.rerun()
            else:
                st.warning("No se encontraron registros válidos para importar.")
                
        except Exception as e:
            st.error(f"Error al importar: {e}")
            import traceback
            st.code(traceback.format_exc())

# Mostrar facturas importadas
if "facturas_importadas" in st.session_state and st.session_state.facturas_importadas:
    st.write(f"**📑 Facturas/Movimientos importados: {len(st.session_state.facturas_importadas)} registros**")
    
    # Resumen por división y tipo
    df_resumen = pd.DataFrame(st.session_state.facturas_importadas)
    resumen = df_resumen.groupby(["Division", "Tipo Movimiento", "Movimiento"]).agg({
        "Cantidad": ["count", "sum"]
    }).reset_index()
    resumen.columns = ["División", "Tipo Mov", "I/E", "Registros", "Total TN"]
    st.dataframe(resumen, use_container_width=True)
    
    if st.button("🗑️ Limpiar facturas importadas"):
        st.session_state.facturas_importadas = []
        st.rerun()

# ===== PASO 2.6: CUADRE CON PROGRAMACIÓN =====
st.subheader("📊 Fase 5: Cuadre con Programación AS 2025")

with st.expander("🔍 Validar datos vs Programación", expanded=False):
    st.markdown("""
    **Este módulo compara los datos extraídos de las divisiones con el Excel de Programación AS 2025.**
    
    📌 **Correspondencias:**
    - **SALV** → Potrerillos + Barquito
    - **TTE** → Caletones + TBA (Los Lirios) + San Antonio  
    - **DMH** → Ministro Hales
    - **CHU** → Chuquicamata + RT + GM
    - **VENT** → Ventana
    - **MEJ** → Mejillones Puerto
    - **MEJICL** → Mejillones Interacid
    - **TERMEJ** → Terquim Mejillones
    """)
    
    archivo_progra = st.file_uploader(
        "Sube el Excel PROGRAMACIÓN AS 2025",
        type=["xlsx"],
        key="progra_import"
    )
    
    # Seleccionar mes a comparar
    mes_progra = st.selectbox(
        "Mes a comparar",
        ["enero", "febrero", "marzo", "abril", "mayo", "junio", 
         "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre"],
        index=10,  # Default noviembre
        key="mes_progra"
    )
    
    # Selector de hojas a comparar
    hojas_comparar = st.multiselect(
        "Selecciona las hojas a extraer",
        ["SALV", "TTE", "DMH", "CHU", "VENT", "MEJ", "MEJICL", "TERMEJ", "COMPRAS"],
        default=["SALV", "TTE", "DMH", "CHU", "VENT"],
        key="hojas_comparar"
    )
    
    if archivo_progra and st.button("📊 Extraer y Comparar", key="btn_extraer_progra"):
        try:
            import openpyxl
            wb_progra = openpyxl.load_workbook(archivo_progra, data_only=True)
            hojas_excel = wb_progra.sheetnames
            st.write(f"📋 Hojas encontradas: {hojas_excel}")
            
            # Obtener columna del mes (D=4=ENE hasta O=15=DIC)
            col_mes = COLUMNA_MES_D_START.get(mes_progra.lower(), 13)  # Default noviembre
            st.info(f"📅 Leyendo columna {col_mes} para {mes_progra.upper()}")
            
            datos_progra = {}
            
            # Función auxiliar para leer celda con manejo de None
            def leer_celda(ws, row, col):
                val = ws.cell(row=row, column=col).value
                return float(val) if val is not None else 0.0
            
            # ========== EXTRAER DATOS DE HOJA "SALV" ==========
            if "SALV" in hojas_comparar and "SALV" in hojas_excel:
                ws = wb_progra["SALV"]
                config = CONFIG_SALV["secciones"]
                
                datos_progra["Potrerillos"] = {
                    "inventario_inicial": leer_celda(ws, config["Potrerillos"]["inventario_inicial"], col_mes),
                    "produccion": leer_celda(ws, config["Potrerillos"]["produccion"], col_mes),
                    "consumo_interno": leer_celda(ws, config["Potrerillos"]["consumo_interno"], col_mes),
                    "excedente": leer_celda(ws, config["Potrerillos"]["excedente"], col_mes),
                    "de_canjes": leer_celda(ws, config["Potrerillos"]["de_canjes"], col_mes),
                    "total_compromisos": leer_celda(ws, config["Potrerillos"]["total_compromisos"], col_mes),
                    "ventas": leer_celda(ws, config["Potrerillos"]["ventas"], col_mes),
                    "canjes_devoluciones": leer_celda(ws, config["Potrerillos"]["canjes_devoluciones"], col_mes),
                }
                
                datos_progra["Barquito"] = {
                    "inventario_inicial": leer_celda(ws, config["Barquito"]["inventario_inicial"], col_mes),
                    "compras": leer_celda(ws, config["Barquito"]["compras"], col_mes),
                    "total_compromisos": leer_celda(ws, config["Barquito"]["total_compromisos"], col_mes),
                    "ventas": leer_celda(ws, config["Barquito"]["ventas"], col_mes),
                    "a_canjes_devoluciones": leer_celda(ws, config["Barquito"]["a_canjes_devoluciones"], col_mes),
                }
                st.success("✅ SALV extraído")
            
            # ========== EXTRAER DATOS DE HOJA "TTE" ==========
            if "TTE" in hojas_comparar and "TTE" in hojas_excel:
                ws = wb_progra["TTE"]
                config = CONFIG_TTE["secciones"]
                
                datos_progra["Caletones"] = {
                    "inventario_inicial": leer_celda(ws, config["Caletones"]["inventario_inicial"], col_mes),
                    "produccion": leer_celda(ws, config["Caletones"]["produccion"], col_mes),
                    "consumo": leer_celda(ws, config["Caletones"]["consumo"], col_mes),
                    "excedentes": leer_celda(ws, config["Caletones"]["excedentes"], col_mes),
                    "traspaso_a_los_lirios": leer_celda(ws, config["Caletones"]["traspaso_a_los_lirios"], col_mes),
                    "saldo_mes": leer_celda(ws, config["Caletones"]["saldo_mes"], col_mes),
                }
                
                datos_progra["Los_Lirios"] = {
                    "inventario_inicial": leer_celda(ws, config["Los_Lirios"]["inventario_inicial"], col_mes),
                    "recepcion_desde_caletones": leer_celda(ws, config["Los_Lirios"]["recepcion_desde_caletones"], col_mes),
                    "ventas": leer_celda(ws, config["Los_Lirios"]["ventas"], col_mes),
                    "canjes_devoluciones": leer_celda(ws, config["Los_Lirios"]["canjes_devoluciones"], col_mes),
                    "traspaso_a_san_antonio": leer_celda(ws, config["Los_Lirios"]["traspaso_a_san_antonio"], col_mes),
                    "saldo_mes": leer_celda(ws, config["Los_Lirios"]["saldo_mes"], col_mes),
                }
                
                datos_progra["San_Antonio"] = {
                    "inventario_inicial": leer_celda(ws, config["San_Antonio"]["inventario_inicial"], col_mes),
                    "recepcion_desde_los_lirios": leer_celda(ws, config["San_Antonio"]["recepcion_desde_los_lirios"], col_mes),
                    "total_compromisos": leer_celda(ws, config["San_Antonio"]["total_compromisos"], col_mes),
                    "ventas": leer_celda(ws, config["San_Antonio"]["ventas"], col_mes),
                    "canjes_devoluciones": leer_celda(ws, config["San_Antonio"]["canjes_devoluciones"], col_mes),
                    "saldo_mes": leer_celda(ws, config["San_Antonio"]["saldo_mes"], col_mes),
                }
                st.success("✅ TTE extraído")
            
            # ========== EXTRAER DATOS DE HOJA "DMH" ==========
            if "DMH" in hojas_comparar and "DMH" in hojas_excel:
                ws = wb_progra["DMH"]
                config = CONFIG_DMH_PROGRA["secciones"]
                
                datos_progra["Ministro_Hales"] = {
                    "inventario_inicial": leer_celda(ws, config["DMH"]["inventario_inicial"], col_mes),
                    "produccion": leer_celda(ws, config["DMH"]["produccion"], col_mes),
                    "consumo": leer_celda(ws, config["DMH"]["consumo"], col_mes),
                    "excedentes": leer_celda(ws, config["DMH"]["excedentes"], col_mes),
                    "compras": leer_celda(ws, config["DMH"]["compras"], col_mes),
                    "retornos_canjes": leer_celda(ws, config["DMH"]["retornos_canjes"], col_mes),
                    "total_compromisos": leer_celda(ws, config["DMH"]["total_compromisos"], col_mes),
                    "ventas": leer_celda(ws, config["DMH"]["ventas"], col_mes),
                    "canjes_devoluciones": leer_celda(ws, config["DMH"]["canjes_devoluciones"], col_mes),
                    "saldo_mes": leer_celda(ws, config["DMH"]["saldo_mes"], col_mes),
                }
                st.success("✅ DMH extraído")
            
            # ========== EXTRAER DATOS DE HOJA "CHU" ==========
            if "CHU" in hojas_comparar and "CHU" in hojas_excel:
                ws = wb_progra["CHU"]
                config = CONFIG_CHUQUI["secciones"]
                
                datos_progra["Chuquicamata"] = {
                    "inventario_inicial": leer_celda(ws, config["Chuquicamata"]["inventario_inicial"], col_mes),
                    "produccion": leer_celda(ws, config["Chuquicamata"]["produccion"], col_mes),
                    "consumo_interno": leer_celda(ws, config["Chuquicamata"]["consumo_interno"], col_mes),
                    "excedentes": leer_celda(ws, config["Chuquicamata"]["excedentes"], col_mes),
                    "compras": leer_celda(ws, config["Chuquicamata"]["compras"], col_mes),
                    "retornos_canjes": leer_celda(ws, config["Chuquicamata"]["retornos_canjes"], col_mes),
                    "total_compromisos": leer_celda(ws, config["Chuquicamata"]["total_compromisos"], col_mes),
                    "ventas": leer_celda(ws, config["Chuquicamata"]["ventas"], col_mes),
                    "canjes_devoluciones": leer_celda(ws, config["Chuquicamata"]["canjes_devoluciones"], col_mes),
                    "saldo_mes": leer_celda(ws, config["Chuquicamata"]["saldo_mes"], col_mes),
                }
                st.success("✅ CHU extraído")
            
            # ========== EXTRAER DATOS DE HOJA "VENT" ==========
            if "VENT" in hojas_comparar and "VENT" in hojas_excel:
                ws = wb_progra["VENT"]
                config = CONFIG_VENT["secciones"]
                
                datos_progra["Ventanas"] = {
                    "inventario_inicial": leer_celda(ws, config["Ventanas"]["inventario_inicial"], col_mes),
                    "produccion": leer_celda(ws, config["Ventanas"]["produccion"], col_mes),
                    "consumo_interno": leer_celda(ws, config["Ventanas"]["consumo_interno"], col_mes),
                    "excedentes": leer_celda(ws, config["Ventanas"]["excedentes"], col_mes),
                    "compras": leer_celda(ws, config["Ventanas"]["compras"], col_mes),
                    "de_canjes": leer_celda(ws, config["Ventanas"]["de_canjes"], col_mes),
                    "total_compromisos": leer_celda(ws, config["Ventanas"]["total_compromisos"], col_mes),
                    "ventas": leer_celda(ws, config["Ventanas"]["ventas"], col_mes),
                    "a_canjes": leer_celda(ws, config["Ventanas"]["a_canjes"], col_mes),
                    "saldo_mensual": leer_celda(ws, config["Ventanas"]["saldo_mensual"], col_mes),
                }
                st.success("✅ VENT extraído")
            
            wb_progra.close()
            
            # Guardar datos extraídos en session_state
            st.session_state.datos_progra = datos_progra
            st.session_state.mes_progra = mes_progra
            
            # Mostrar datos extraídos
            st.success(f"✅ Datos de Programación extraídos para {mes_progra.upper()}")
            
            for grupo, valores in datos_progra.items():
                with st.expander(f"📁 {grupo}", expanded=False):
                    df_grupo = pd.DataFrame([
                        {"Concepto": k.replace("_", " ").title(), "Valor (TM)": f"{v:,.2f}"}
                        for k, v in valores.items() if v != 0
                    ])
                    st.dataframe(df_grupo, use_container_width=True)
            
        except Exception as e:
            st.error(f"Error al extraer datos de Programación: {e}")
            import traceback
            st.code(traceback.format_exc())

# Mostrar comparación si hay datos de ambas fuentes
if "datos_progra" in st.session_state and "df_completo" in st.session_state:
    st.write("---")
    st.subheader("📈 Comparación: Programación vs Divisiones")
    
    df_completo = st.session_state.df_completo
    datos_progra = st.session_state.datos_progra
    mes_progra = st.session_state.get("mes_progra", "")
    
    st.info(f"📅 Comparando datos de **{mes_progra.upper()}**")
    
    # Agregar datos de divisiones por grupo
    comparaciones = []
    
    # ===== POTRERILLOS =====
    if "Potrerillos" in datos_progra:
        df_div = df_completo[df_completo["Division"] == "Potrerillos"]
        
        # Producción
        prod_div = df_div[df_div["Tipo Movimiento"] == "MPRO"]["Cantidad"].sum()
        comparaciones.append({
            "División": "Potrerillos", "Concepto": "Producción",
            "Programación": datos_progra["Potrerillos"].get("produccion", 0),
            "Balance": prod_div,
        })
        
        # Ventas
        ventas_div = df_div[df_div["Tipo Movimiento"] == "VENT"]["Cantidad"].sum()
        comparaciones.append({
            "División": "Potrerillos", "Concepto": "Ventas",
            "Programación": datos_progra["Potrerillos"].get("ventas", 0),
            "Balance": ventas_div,
        })
    
    # ===== BARQUITO =====
    if "Barquito" in datos_progra:
        df_div = df_completo[df_completo["Division"] == "Barquito"]
        
        # Ventas
        ventas_div = df_div[df_div["Tipo Movimiento"] == "VENT"]["Cantidad"].sum()
        comparaciones.append({
            "División": "Barquito", "Concepto": "Ventas",
            "Programación": datos_progra["Barquito"].get("ventas", 0),
            "Balance": ventas_div,
        })
    
    # ===== CALETONES =====
    if "Caletones" in datos_progra:
        df_div = df_completo[df_completo["Division"] == "Caletones"]
        
        # Producción
        prod_div = df_div[df_div["Tipo Movimiento"] == "MPRO"]["Cantidad"].sum()
        comparaciones.append({
            "División": "Caletones", "Concepto": "Producción",
            "Programación": datos_progra["Caletones"].get("produccion", 0),
            "Balance": prod_div,
        })
        
        # Consumo
        cons_div = df_div[df_div["Tipo Movimiento"] == "ECIP"]["Cantidad"].sum()
        comparaciones.append({
            "División": "Caletones", "Concepto": "Consumo",
            "Programación": datos_progra["Caletones"].get("consumo", 0),
            "Balance": cons_div,
        })
    
    # ===== LOS LIRIOS (TBA) =====
    if "Los_Lirios" in datos_progra:
        df_div = df_completo[df_completo["Division"] == "TBA"]
        
        # Ventas
        ventas_div = df_div[df_div["Tipo Movimiento"] == "VENT"]["Cantidad"].sum()
        comparaciones.append({
            "División": "Los Lirios (TBA)", "Concepto": "Ventas",
            "Programación": datos_progra["Los_Lirios"].get("ventas", 0),
            "Balance": ventas_div,
        })
    
    # ===== SAN ANTONIO =====
    if "San_Antonio" in datos_progra:
        df_div = df_completo[df_completo["Division"] == "San Antonio"]
        
        # Ventas
        ventas_div = df_div[df_div["Tipo Movimiento"] == "VENT"]["Cantidad"].sum()
        comparaciones.append({
            "División": "San Antonio", "Concepto": "Ventas",
            "Programación": datos_progra["San_Antonio"].get("ventas", 0),
            "Balance": ventas_div,
        })
    
    # ===== MINISTRO HALES =====
    if "Ministro_Hales" in datos_progra:
        df_div = df_completo[df_completo["Division"] == "Ministro Hales"]
        
        # Producción
        prod_div = df_div[df_div["Tipo Movimiento"] == "MPRO"]["Cantidad"].sum()
        comparaciones.append({
            "División": "Ministro Hales", "Concepto": "Producción",
            "Programación": datos_progra["Ministro_Hales"].get("produccion", 0),
            "Balance": prod_div,
        })
        
        # Ventas
        ventas_div = df_div[df_div["Tipo Movimiento"] == "VENT"]["Cantidad"].sum()
        comparaciones.append({
            "División": "Ministro Hales", "Concepto": "Ventas",
            "Programación": datos_progra["Ministro_Hales"].get("ventas", 0),
            "Balance": ventas_div,
        })
    
    # ===== CHUQUICAMATA =====
    if "Chuquicamata" in datos_progra:
        df_div = df_completo[df_completo["Division"] == "Chuquicamata"]
        
        # Producción
        prod_div = df_div[df_div["Tipo Movimiento"] == "MPRO"]["Cantidad"].sum()
        comparaciones.append({
            "División": "Chuquicamata", "Concepto": "Producción",
            "Programación": datos_progra["Chuquicamata"].get("produccion", 0),
            "Balance": prod_div,
        })
        
        # Ventas
        ventas_div = df_div[df_div["Tipo Movimiento"] == "VENT"]["Cantidad"].sum()
        comparaciones.append({
            "División": "Chuquicamata", "Concepto": "Ventas",
            "Programación": datos_progra["Chuquicamata"].get("ventas", 0),
            "Balance": ventas_div,
        })
    
    # ===== VENTANAS =====
    if "Ventanas" in datos_progra:
        df_div = df_completo[df_completo["Division"] == "Ventana"]
        
        # Producción
        prod_div = df_div[df_div["Tipo Movimiento"] == "MPRO"]["Cantidad"].sum()
        comparaciones.append({
            "División": "Ventanas", "Concepto": "Producción",
            "Programación": datos_progra["Ventanas"].get("produccion", 0),
            "Balance": prod_div,
        })
        
        # Ventas
        ventas_div = df_div[df_div["Tipo Movimiento"] == "VENT"]["Cantidad"].sum()
        comparaciones.append({
            "División": "Ventanas", "Concepto": "Ventas",
            "Programación": datos_progra["Ventanas"].get("ventas", 0),
            "Balance": ventas_div,
        })
    
    if comparaciones:
        df_comp = pd.DataFrame(comparaciones)
        df_comp["Diferencia"] = df_comp["Programación"] - df_comp["Balance"]
        df_comp["% Dif"] = df_comp.apply(
            lambda row: f"{(row['Diferencia']/row['Programación']*100):.1f}%" if row['Programación'] != 0 else "N/A",
            axis=1
        )
        df_comp["Estado"] = df_comp["Diferencia"].apply(
            lambda x: "✅" if abs(x) <= TOLERANCIA_DIFERENCIA else ("⚠️" if abs(x) <= 100 else "❌")
        )
        
        # Formatear números
        df_display = df_comp.copy()
        df_display["Programación"] = df_display["Programación"].apply(lambda x: f"{x:,.2f}")
        df_display["Balance"] = df_display["Balance"].apply(lambda x: f"{x:,.2f}")
        df_display["Diferencia"] = df_display["Diferencia"].apply(lambda x: f"{x:,.2f}")
        
        st.dataframe(df_display, use_container_width=True, hide_index=True)
        
        # Resumen
        ok_count = len(df_comp[df_comp["Estado"] == "✅"])
        warn_count = len(df_comp[df_comp["Estado"] == "⚠️"])
        error_count = len(df_comp[df_comp["Estado"] == "❌"])
        
        col1, col2, col3 = st.columns(3)
        col1.metric("✅ Cuadran", ok_count)
        col2.metric("⚠️ Diferencia pequeña", warn_count)
        col3.metric("❌ No cuadran", error_count)
    else:
        st.info("No hay datos suficientes para comparar. Procesa los archivos de Balance primero.")

# ===== PASO 3: AJUSTES MANUALES POR DIVISIÓN =====
st.subheader("🔧 Fase 4: Ajustes Manuales (Opcional)")

# Lista de divisiones disponibles para ajustes
DIVISIONES_AJUSTES = ["Salvador", "Caletones", "TBA", "San Antonio", "Potrerillos", "Chuquicamata", 
                      "Radomiro Tomic", "Gabriela Mistral", "Ministro Hales", "Barquito"]

# Códigos de bodega por división
BODEGAS_DIVISION = {
    "Salvador": "SALVA",
    "Caletones": "TECA",
    "TBA": "TETBA", 
    "San Antonio": "CM",
    "Potrerillos": "SALPO",
    "Chuquicamata": "CH",
    "Radomiro Tomic": "RT",
    "Gabriela Mistral": "DGM",
    "Ministro Hales": "CM",
    "Barquito": "BARQUITO",
}

# Tipos de movimiento disponibles
TIPOS_MOVIMIENTO = ["MDEV", "EDEV", "MPRO", "ECIP", "TIEB", "TIPB", "VENT", "AJUSTE"]

# Inicializar lista de ajustes en session_state
if "ajustes_manuales" not in st.session_state:
    st.session_state.ajustes_manuales = []

with st.expander("➕ Agregar Ajuste Manual", expanded=False):
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        ajuste_division = st.selectbox("División", DIVISIONES_AJUSTES, key="ajuste_div")
        ajuste_bodega = st.text_input("Id. Bodega", value=BODEGAS_DIVISION.get(DIVISIONES_AJUSTES[0], ""), key="ajuste_bodega")
        ajuste_rut = st.text_input("Rut", value="61.704.000-K", key="ajuste_rut")
    
    with col2:
        ajuste_num_factura = st.text_input("Número Factura", key="ajuste_factura")
        ajuste_tipo_mov = st.selectbox("Tipo Movimiento", TIPOS_MOVIMIENTO, key="ajuste_tipo")
        ajuste_nombre = st.text_input("Nombre", value="Corporación Nacional del Cobre de Chile (Codelco)", key="ajuste_nombre")
    
    with col3:
        ajuste_movimiento = st.selectbox("Ingreso/Egreso", ["I", "E"], key="ajuste_ie")
        ajuste_cantidad = st.number_input("Cantidad (TN)", min_value=0.0, step=0.001, format="%.3f", key="ajuste_cant")
        ajuste_expediente = st.text_input("Expediente Com. Anticipada", key="ajuste_expediente")
    
    with col4:
        ajuste_material = st.text_input("SQC CAS/Nombre Mezcla", value="7664-93-9", key="ajuste_material")
        ajuste_concentracion = st.text_input("Concentración", value="Ácido Sulfúrico (Conc: 96)", key="ajuste_conc")
        ajuste_unidad = st.selectbox("Unidad Medida", ["TN", "KG", "LT"], key="ajuste_unidad")
    
    if st.button("➕ Agregar Ajuste"):
        if ajuste_cantidad > 0:
            mes_nombre = fecha.strftime("%b").upper()  # NOV, DIC, etc.
            nuevo_ajuste = {
                "Division": ajuste_division,
                "Concepto": ajuste_num_factura if ajuste_num_factura else f"K.{ajuste_movimiento}.Ajuste.{mes_nombre}",
                "Grupo": ajuste_num_factura if ajuste_num_factura else f"ajuste_{ajuste_division.lower()}_{len(st.session_state.ajustes_manuales)}",
                "Fecha": fecha,
                "Cantidad": ajuste_cantidad,
                "Unidad": ajuste_unidad,
                "IncludeBatch": True,
                "Bodega": ajuste_bodega if ajuste_bodega else BODEGAS_DIVISION.get(ajuste_division, ""),
                "Material": ajuste_material,
                "Tipo Movimiento": ajuste_tipo_mov,
                "Movimiento": ajuste_movimiento,
                "Concentracion": ajuste_concentracion,
                "Expediente": ajuste_expediente,
                "Rut": ajuste_rut,
                "Nombre": ajuste_nombre,
            }
            st.session_state.ajustes_manuales.append(nuevo_ajuste)
            
            # Si ya se procesaron datos, actualizar df_completo con el nuevo ajuste
            if "df_base" in st.session_state and st.session_state.df_base is not None:
                df_ajustes = pd.DataFrame(st.session_state.ajustes_manuales)
                if st.session_state.df_base.empty:
                    st.session_state.df_completo = df_ajustes.copy()
                else:
                    st.session_state.df_completo = pd.concat([st.session_state.df_base, df_ajustes], ignore_index=True)
            
            st.success(f"✅ Ajuste agregado: {ajuste_division} - {ajuste_tipo_mov} {ajuste_movimiento} {ajuste_cantidad} TN")
            st.rerun()

# Mostrar ajustes agregados
if st.session_state.ajustes_manuales:
    st.write("**Ajustes agregados:**")
    for i, ajuste in enumerate(st.session_state.ajustes_manuales):
        col1, col2 = st.columns([4, 1])
        with col1:
            factura_txt = f"[{ajuste.get('Concepto', '')}] " if ajuste.get('Concepto') else ""
            st.write(f"• {ajuste['Division']}: {factura_txt}{ajuste['Tipo Movimiento']} {ajuste['Movimiento']} {ajuste['Cantidad']:.3f} {ajuste.get('Unidad', 'TN')} - {ajuste.get('Rut', '')} {ajuste.get('Nombre', '')[:30]}...")
        with col2:
            if st.button("🗑️", key=f"del_ajuste_{i}"):
                st.session_state.ajustes_manuales.pop(i)
                # También actualizar df_completo al eliminar
                if "df_base" in st.session_state and st.session_state.df_base is not None:
                    if st.session_state.ajustes_manuales:
                        df_ajustes = pd.DataFrame(st.session_state.ajustes_manuales)
                        st.session_state.df_completo = pd.concat([st.session_state.df_base, df_ajustes], ignore_index=True)
                    else:
                        st.session_state.df_completo = st.session_state.df_base.copy()
                st.rerun()
    
    if st.button("🗑️ Limpiar todos los ajustes"):
        st.session_state.ajustes_manuales = []
        if "df_base" in st.session_state and st.session_state.df_base is not None:
            st.session_state.df_completo = st.session_state.df_base.copy()
        st.rerun()
    
    # Mensaje informativo
    if "df_base" in st.session_state:
        st.info("💡 Los ajustes se actualizarán automáticamente en los datos procesados.")

# Verificar si hay facturas importadas
facturas_importadas = st.session_state.get("facturas_importadas", [])

# Botón para procesar
if st.button("🚀 Cargar y Procesar", type="primary", use_container_width=True):
    # Verificar que haya archivos O ajustes manuales O facturas importadas
    if not archivos and not st.session_state.ajustes_manuales and not facturas_importadas:
        st.error("Por favor sube al menos un archivo de Balance, agrega ajustes manuales o importa facturas")
        st.stop()
    
    todos_los_datos = []
    
    # Procesar archivos si existen
    if archivos:
        # Mostrar archivos detectados
        st.subheader("📋 Archivos cargados")
        archivos_por_division = {}
        
        for archivo in archivos:
            division = detectar_division(archivo.name)
            if division:
                # Si es Caletones, agregar también TBA y San Antonio (mismo Excel, diferentes columnas)
                if division == "Caletones":
                    archivos_por_division["Caletones"] = archivo
                    archivos_por_division["TBA"] = archivo
                    archivos_por_division["San Antonio"] = archivo
                    st.success(f"✅ {archivo.name} → **Caletones, TBA y San Antonio**")
                else:
                    archivos_por_division[division] = archivo
                    st.success(f"✅ {archivo.name} → **{division}**")
            else:
                st.warning(f"⚠️ {archivo.name} → No se pudo detectar la división")
    
        if not archivos_por_division:
            st.error("No se detectó ninguna división. Verifica los nombres de los archivos.")
            if not st.session_state.ajustes_manuales:
                st.stop()
            else:
                st.info("📝 Continuando solo con ajustes manuales...")
        
        # Extraer datos de todas las divisiones
        with st.spinner("Leyendo archivos…"):
            for division, archivo in archivos_por_division.items():
                try:
                    config = DIVISIONES_CONFIG.get(division)
                    if not config:
                        st.warning(f"⚠️ No hay configuración para {division}")
                        continue
                    
                    st.info(f"🔍 Procesando {division} desde {archivo.name}...")
                        
                    df = extraer_movimientos(
                        archivo,
                        config,
                        fecha,
                        division
                    )
                    if not df.empty:
                        todos_los_datos.append(df)
                        st.success(f"✅ {division}: {len(df)} registros extraídos")
                    else:
                        st.warning(f"⚠️ {division}: No se encontraron datos (celdas vacías o con valor 0)")
                except Exception as exc:
                    st.error(f"❌ Error en {division} ({archivo.name}): {exc}")
                    import traceback
                    st.code(traceback.format_exc())
    else:
        if st.session_state.ajustes_manuales or facturas_importadas:
            st.info("📝 Procesando solo con ajustes manuales/facturas importadas (sin archivos de Balance)")
    
    # Combinar todos los DataFrames (puede estar vacío si solo hay ajustes)
    if todos_los_datos:
        df_completo = pd.concat(todos_los_datos, ignore_index=True)
    else:
        df_completo = pd.DataFrame()
    
    # Guardar el DataFrame base (sin ajustes ni facturas) en session_state
    st.session_state.df_base = df_completo.copy() if not df_completo.empty else pd.DataFrame()
    
    # Agregar facturas importadas si existen
    if facturas_importadas:
        df_facturas = pd.DataFrame(facturas_importadas)
        if df_completo.empty:
            df_completo = df_facturas.copy()
        else:
            df_completo = pd.concat([df_completo, df_facturas], ignore_index=True)
        st.info(f"📑 Se agregaron {len(facturas_importadas)} factura(s) importada(s)")
    
    # Agregar ajustes manuales si existen
    if st.session_state.ajustes_manuales:
        df_ajustes = pd.DataFrame(st.session_state.ajustes_manuales)
        if df_completo.empty:
            df_completo = df_ajustes.copy()
        else:
            df_completo = pd.concat([df_completo, df_ajustes], ignore_index=True)
        st.info(f"📝 Se agregaron {len(st.session_state.ajustes_manuales)} ajuste(s) manual(es)")
    
    # Guardar df_completo con ajustes en session_state para persistencia
    st.session_state.df_completo = df_completo.copy()
    
    if df_completo.empty:
        st.warning("No se encontraron movimientos en los archivos.")
        st.stop()

    st.subheader("📊 Datos extraídos")
    st.dataframe(df_completo, use_container_width=True)
    
    # Cargar SIREGAD si se proporcionó
    df_siregad = None
    if archivo_siregad:
        try:
            with st.spinner("Leyendo respaldo SIREGAD…"):
                df_siregad = extraer_siregad(archivo_siregad)
            st.success(f"✅ Respaldo SIREGAD cargado: {len(df_siregad)} registros")
        except Exception as exc:
            st.warning(f"⚠️ No se pudo leer el respaldo SIREGAD: {exc}")
            import traceback
            st.code(traceback.format_exc())
    
    # Botón para descargar Excel intermedio con formato
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from config import RUT_EMPRESAS
    
    buffer_intermedio = io.BytesIO()
    
    with pd.ExcelWriter(buffer_intermedio, engine='openpyxl') as writer:
        # Crear la hoja "Datos" una sola vez
        df_dummy = pd.DataFrame()
        df_dummy.to_excel(writer, sheet_name='Datos', index=False)
        ws = writer.sheets['Datos']
        
        row_actual = 1
        
        # Colores por división - Verde manzana pastel para todas
        color_verde_manzana = "C1E1C1"  # Verde manzana pastel
        colores_division = {
            "Salvador": color_verde_manzana,
            "Caletones": color_verde_manzana,
            "Potrerillos": color_verde_manzana,
            "TBA": color_verde_manzana,
            "San Antonio": color_verde_manzana,
            "Chuquicamata": color_verde_manzana,
            "Radomiro Tomic": color_verde_manzana,
            "Gabriela Mistral": color_verde_manzana,
            "Ministro Hales": color_verde_manzana,
            "Barquito": color_verde_manzana,
        }
        
        # Header de columnas según formato SIREGAD
        columnas = ["dd-mm-aaaa", "Número Factura", "Id. Bodega", "SQC CAS/Nombre Mezcla", 
                   "Concentración (0,00000)", "Ingreso/Egreso", "Tipo de Movimiento", "Cantidad", 
                   "Unidad Medida", "Expediente Comunicación Anticipada", "Rut", "Nombre"]
        
        # Orden de divisiones según especificación
        orden_divisiones = ["Potrerillos", "Caletones", "TBA", "Radomiro Tomic", 
                          "Chuquicamata", "Gabriela Mistral", "Ministro Hales", "Barquito", 
                          "San Antonio"]
        
        # Filtrar divisiones que existen en df_completo y mantener el orden
        divisiones_ordenadas = [div for div in orden_divisiones if div in df_completo["Division"].unique()]
        
        # ===== HEADER GLOBAL (UNA SOLA VEZ AL INICIO) =====
        for col_idx, col_name in enumerate(columnas, 1):
            cell = ws.cell(row=row_actual, column=col_idx)
            cell.value = col_name
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        row_actual += 1
        
        for division in divisiones_ordenadas:
            df_div = df_completo[df_completo["Division"] == division].copy()
            
            # Calcular totales simplificados
            inv_inicial = df_div[df_div["Concepto"].str.contains("inventario_inicial", na=False)]["Cantidad"].sum()
            ingresos_total = df_div[df_div["Movimiento"] == "I"]["Cantidad"].sum()
            egresos_total = df_div[df_div["Movimiento"] == "E"]["Cantidad"].sum()
            inv_final_calculado = inv_inicial + ingresos_total - egresos_total
            
            # Extraer el inventario final del Excel
            inv_final_extraido = df_div[df_div["Concepto"].str.contains("inventario_final", na=False)]["Cantidad"].sum()
            
            # Calcular diferencia
            diferencia = inv_final_extraido - inv_final_calculado
            
            color = colores_division.get(division, "FFFFFF")
            
            # ===== HEADER DE DIVISIÓN =====
            # Escribir valores ANTES de fusionar
            cell_header = ws[f'A{row_actual}']
            cell_header.value = division.upper()
            cell_header.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            cell_header.font = Font(bold=True, size=12)
            cell_header.alignment = Alignment(horizontal="left", vertical="center")
            ws.merge_cells(f'A{row_actual}:G{row_actual}')
            
            # Agregar INV INICIAL en columna H (donde van las cantidades)
            cell_inv_ini_valor = ws[f'H{row_actual}']
            cell_inv_ini_valor.value = inv_inicial
            cell_inv_ini_valor.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            cell_inv_ini_valor.font = Font(bold=True, size=11)
            cell_inv_ini_valor.alignment = Alignment(horizontal="center")
            cell_inv_ini_valor.number_format = '#,##0.000'
            
            row_actual += 1
            
            # ===== DATOS DE MOVIMIENTOS =====
            # Forzar conversión de IncludeBatch a booleano
            df_div["IncludeBatch"] = df_div["IncludeBatch"].astype(bool)
            df_batch = df_div[df_div["IncludeBatch"] == True].copy()
            
            # Agrupar por "Grupo" (suma celdas con mismo grupo), sumando las cantidades
            # Primero asegurar que existan las columnas adicionales
            for col in ["Concentracion", "Expediente", "Rut", "Nombre"]:
                if col not in df_batch.columns:
                    df_batch[col] = ""
            
            df_batch_agrupado = df_batch.groupby("Grupo", as_index=False).agg({
                "Cantidad": "sum",
                "Bodega": "first",
                "Material": "first",
                "Tipo Movimiento": "first",
                "Movimiento": "first",
                "Unidad": "first",
                "Concentracion": "first",
                "Expediente": "first",
                "Rut": "first",
                "Nombre": "first"
            }).copy()
            
            for idx, row in df_batch_agrupado.iterrows():
                # Columna 1: dd-mm-aaaa (Fecha)
                ws.cell(row=row_actual, column=1).value = fecha_ultimo_dia
                
                # Columna 2: Número Factura - Generar código especial para NO VENT
                movimiento = row.get("Movimiento", "")
                tipo_mov = row.get("Tipo Movimiento", "")
                grupo = row.get("Grupo", "")
                
                # Si NO es "E VENT", generar el código
                if not (movimiento == "E" and tipo_mov == "VENT"):
                    # Si es ajuste (manual o inventario), usar 'Ajuste' en lugar del tipo_mov
                    if "ajuste" in grupo.lower():
                        numero_factura = f"'K.{movimiento}.Ajuste.DIC'"
                    else:
                        numero_factura = f"'K.{movimiento}.{tipo_mov}.DIC'"
                else:
                    # Para E VENT, usar el grupo (número de factura real)
                    numero_factura = f"'{grupo}'"
                
                ws.cell(row=row_actual, column=2).value = numero_factura
                
                # Columna 3: Id. Bodega
                ws.cell(row=row_actual, column=3).value = row.get("Bodega", "")
                # Columna 4: SQC CAS/Nombre Mezcla
                ws.cell(row=row_actual, column=4).value = row.get("Material", "")
                # Columna 5: Concentración (0,00000)
                conc = row.get("Concentracion", "")
                ws.cell(row=row_actual, column=5).value = conc if conc else "Ácido Sulfúrico (Conc: 96)"
                # Columna 6: Ingreso/Egreso
                ws.cell(row=row_actual, column=6).value = row.get("Movimiento", "")
                # Columna 7: Tipo de Movimiento
                ws.cell(row=row_actual, column=7).value = row.get("Tipo Movimiento", "")
                # Columna 8: Cantidad
                ws.cell(row=row_actual, column=8).value = row.get("Cantidad", 0)
                # Columna 9: Unidad Medida
                ws.cell(row=row_actual, column=9).value = row.get("Unidad", "TN")
                # Columna 10: Expediente Comunicación Anticipada
                ws.cell(row=row_actual, column=10).value = row.get("Expediente", "")
                # Columna 11: Rut
                rut = row.get("Rut", "")
                ws.cell(row=row_actual, column=11).value = rut if rut else "61.704.000-K"
                # Columna 12: Nombre
                nombre = row.get("Nombre", "")
                ws.cell(row=row_actual, column=12).value = nombre if nombre else "Corporación Nacional del Cobre de Chile (Codelco)"
                
                # Aplicar color a todas las columnas
                for col in range(1, 13):
                    cell = ws.cell(row=row_actual, column=col)
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                    if col == 8:  # Cantidad
                        cell.number_format = '#,##0.000'
                    if col == 1:  # Fecha
                        cell.number_format = 'DD-MM-YYYY'
                
                row_actual += 1
            
            # ===== TOTALES DE MOVIMIENTOS =====
            # Total de Ingresos (I)
            total_ingresos = df_div[(df_div["Movimiento"] == "I") & (~df_div["Concepto"].str.contains("inventario", na=False))]["Cantidad"].sum()
            
            cell_total_i_label = ws[f'A{row_actual}']
            cell_total_i_label.value = "TOTAL INGRESOS (I)"
            cell_total_i_label.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            cell_total_i_label.font = Font(bold=True)
            cell_total_i_label.alignment = Alignment(horizontal="right")
            ws.merge_cells(f'A{row_actual}:G{row_actual}')
            
            cell_total_i_valor = ws[f'H{row_actual}']
            cell_total_i_valor.value = total_ingresos
            cell_total_i_valor.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            cell_total_i_valor.font = Font(bold=True)
            cell_total_i_valor.alignment = Alignment(horizontal="center")
            cell_total_i_valor.number_format = '#,##0.000'
            
            row_actual += 1
            
            # Total de Egresos (E)
            total_egresos = df_div[(df_div["Movimiento"] == "E") & (~df_div["Concepto"].str.contains("inventario", na=False))]["Cantidad"].sum()
            
            cell_total_e_label = ws[f'A{row_actual}']
            cell_total_e_label.value = "TOTAL EGRESOS (E)"
            cell_total_e_label.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            cell_total_e_label.font = Font(bold=True)
            cell_total_e_label.alignment = Alignment(horizontal="right")
            ws.merge_cells(f'A{row_actual}:G{row_actual}')
            
            cell_total_e_valor = ws[f'H{row_actual}']
            cell_total_e_valor.value = total_egresos
            cell_total_e_valor.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            cell_total_e_valor.font = Font(bold=True)
            cell_total_e_valor.alignment = Alignment(horizontal="center")
            cell_total_e_valor.number_format = '#,##0.000'
            
            row_actual += 1
            
            # Total general (todos los movimientos sin inventarios)
            total_movimientos = df_div[~df_div["Concepto"].str.contains("inventario", na=False)]["Cantidad"].sum()
            
            cell_total_gen_label = ws[f'A{row_actual}']
            cell_total_gen_label.value = "TOTAL MOVIMIENTOS"
            cell_total_gen_label.fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
            cell_total_gen_label.font = Font(bold=True, color="FFFFFF")
            cell_total_gen_label.alignment = Alignment(horizontal="right")
            ws.merge_cells(f'A{row_actual}:G{row_actual}')
            
            cell_total_gen_valor = ws[f'H{row_actual}']
            cell_total_gen_valor.value = total_movimientos
            cell_total_gen_valor.fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
            cell_total_gen_valor.font = Font(bold=True, color="FFFFFF")
            cell_total_gen_valor.alignment = Alignment(horizontal="center")
            cell_total_gen_valor.number_format = '#,##0.000'
            
            row_actual += 1
            
            # ===== FOOTER INV FINAL (en una nueva fila) =====
            # Primero escribir valores, luego fusionar
            
            # INV FINAL CALCULADO - Label
            cell_inv_calc_label = ws[f'A{row_actual}']
            cell_inv_calc_label.value = "INV FINAL (CALCULADO)"
            cell_inv_calc_label.fill = PatternFill(start_color="FF69B4", end_color="FF69B4", fill_type="solid")
            cell_inv_calc_label.font = Font(bold=True)
            cell_inv_calc_label.alignment = Alignment(horizontal="right")
            ws.merge_cells(f'A{row_actual}:B{row_actual}')
            
            # INV FINAL CALCULADO - Valor
            cell_inv_calc_valor = ws[f'C{row_actual}']
            cell_inv_calc_valor.value = inv_final_calculado
            cell_inv_calc_valor.fill = PatternFill(start_color="FF69B4", end_color="FF69B4", fill_type="solid")
            cell_inv_calc_valor.font = Font(bold=True)
            cell_inv_calc_valor.alignment = Alignment(horizontal="right")
            cell_inv_calc_valor.number_format = '#,##0.000'
            ws.merge_cells(f'C{row_actual}:D{row_actual}')
            
            # INV FINAL EXTRAIDO - Label
            cell_inv_ext_label = ws[f'E{row_actual}']
            cell_inv_ext_label.value = "INV FINAL (EXTRAÍDO)"
            cell_inv_ext_label.fill = PatternFill(start_color="FF69B4", end_color="FF69B4", fill_type="solid")
            cell_inv_ext_label.font = Font(bold=True)
            cell_inv_ext_label.alignment = Alignment(horizontal="right")
            ws.merge_cells(f'E{row_actual}:F{row_actual}')
            
            # INV FINAL EXTRAIDO - Valor
            cell_inv_ext_valor = ws[f'G{row_actual}']
            cell_inv_ext_valor.value = inv_final_extraido
            cell_inv_ext_valor.fill = PatternFill(start_color="FF69B4", end_color="FF69B4", fill_type="solid")
            cell_inv_ext_valor.font = Font(bold=True)
            cell_inv_ext_valor.alignment = Alignment(horizontal="right")
            cell_inv_ext_valor.number_format = '#,##0.000'
            ws.merge_cells(f'G{row_actual}:H{row_actual}')
            
            # DIFERENCIA - Label
            cell_dif_label = ws[f'I{row_actual}']
            cell_dif_label.value = "DIFERENCIA"
            cell_dif_label.fill = PatternFill(start_color="FF69B4", end_color="FF69B4", fill_type="solid")
            cell_dif_label.font = Font(bold=True)
            cell_dif_label.alignment = Alignment(horizontal="right")
            
            # DIFERENCIA - Valor
            cell_dif_valor = ws[f'J{row_actual}']
            cell_dif_valor.value = diferencia
            # Color rojo si hay diferencia, verde si cuadra
            color_dif = "FF0000" if abs(diferencia) > 0.001 else "00B050"
            cell_dif_valor.fill = PatternFill(start_color=color_dif, end_color=color_dif, fill_type="solid")
            cell_dif_valor.font = Font(bold=True, color="FFFFFF")
            cell_dif_valor.alignment = Alignment(horizontal="right")
            cell_dif_valor.number_format = '#,##0.000'
            ws.merge_cells(f'J{row_actual}:L{row_actual}')
            
            row_actual += 3  # Espacio entre divisiones
        
        # ===== SECCIÓN TERMINALES MEJILLONES =====
        color_terminales = "FFD700"  # Dorado para terminales
        
        # Header de Terminales Mejillones
        cell_header_term = ws[f'A{row_actual}']
        cell_header_term.value = "TERMINALES MEJILLONES"
        cell_header_term.fill = PatternFill(start_color=color_terminales, end_color=color_terminales, fill_type="solid")
        cell_header_term.font = Font(bold=True, size=12)
        cell_header_term.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(f'A{row_actual}:L{row_actual}')
        row_actual += 1
        
        # Encabezados de columnas
        for col_idx, col_name in enumerate(columnas, 1):
            cell = ws.cell(row=row_actual, column=col_idx)
            cell.value = col_name
            cell.fill = PatternFill(start_color=color_terminales, end_color=color_terminales, fill_type="solid")
            cell.font = Font(bold=True, color="000000")
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        row_actual += 1
        
        # Datos de los 3 terminales
        terminales_mejillones = [
            {"nombre": "Interacid", "bodega": "CM", "cantidad": 64249.702},
            {"nombre": "TPM", "bodega": "CM", "cantidad": 8772.753},
            {"nombre": "Terquim", "bodega": "CM", "cantidad": 5598.250},
        ]
        
        for terminal in terminales_mejillones:
            # Columna 1: Fecha
            ws.cell(row=row_actual, column=1).value = fecha_ultimo_dia
            # Columna 2: Número Factura - Código especial
            ws.cell(row=row_actual, column=2).value = f"K.E.TIEB.DIC"
            # Columna 3: Id. Bodega
            ws.cell(row=row_actual, column=3).value = terminal["bodega"]
            # Columna 4: Material
            ws.cell(row=row_actual, column=4).value = "7664-93-9"
            # Columna 5: Concentración
            ws.cell(row=row_actual, column=5).value = "Ácido Sulfúrico (Conc: 96)"
            # Columna 6: Ingreso/Egreso
            ws.cell(row=row_actual, column=6).value = "E"
            # Columna 7: Tipo de Movimiento
            ws.cell(row=row_actual, column=7).value = "TIEB"
            # Columna 8: Cantidad
            ws.cell(row=row_actual, column=8).value = terminal["cantidad"]
            # Columna 9: Unidad
            ws.cell(row=row_actual, column=9).value = "TN"
            # Columna 10: Expediente
            ws.cell(row=row_actual, column=10).value = ""
            # Columna 11: Rut Codelco
            ws.cell(row=row_actual, column=11).value = "61.704.000-K"
            # Columna 12: Nombre Codelco
            ws.cell(row=row_actual, column=12).value = "Corporación Nacional del Cobre de Chile (Codelco)"
            
            # Aplicar color y formato
            for col in range(1, 13):
                cell = ws.cell(row=row_actual, column=col)
                cell.fill = PatternFill(start_color=color_terminales, end_color=color_terminales, fill_type="solid")
                if col == 8:  # Cantidad
                    cell.number_format = '#,##0.000'
                if col == 1:  # Fecha
                    cell.number_format = 'DD-MM-YYYY'
            
            row_actual += 1
        
        row_actual += 2  # Espacio después de terminales
        
        # Ajustar anchos de columna
        ws = writer.sheets['Datos']
        # Anchos: Fecha, Nº Factura, Bodega, CAS, Concentración, I/E, Tipo Mov, Cantidad, Unidad, Expediente, Rut, Nombre
        anchos = [12, 18, 12, 15, 25, 12, 18, 12, 12, 30, 15, 50]
        for col_idx, ancho in enumerate(anchos, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = ancho
        
        # Agregar pestaña de respaldo SIREGAD si se proporcionó
        if df_siregad is not None and not df_siregad.empty:
            df_siregad.to_excel(writer, sheet_name='Respaldo SIREGAD', index=False)
            ws_respaldo = writer.sheets['Respaldo SIREGAD']
            
            # Formatear headers
            for col_idx in range(1, len(df_siregad.columns) + 1):
                cell = ws_respaldo.cell(row=1, column=col_idx)
                cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", wrap_text=True)
            
            # Ajustar anchos
            for col_idx in range(1, len(df_siregad.columns) + 1):
                ws_respaldo.column_dimensions[get_column_letter(col_idx)].width = 18
    
    buffer_intermedio.seek(0)
    
    st.download_button(
        "📥 Descargar Excel Intermedio",
        buffer_intermedio,
        file_name=f"Intermedio_SIREGAD_{fecha_ultimo_dia.strftime('%Y-%m-%d')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    # Validación
    validacion = validar_inventario(df_completo)

    st.subheader("🔎 Validación por División")
    for resultado in validacion["por_division"]:
        with st.expander(f"División: {resultado['division']}", expanded=True):
            col1, col2, col3, col4 = st.columns(4)
            col1.metric("Inv. Inicial", f"{resultado['inventario_inicial']:.2f} TN")
            col2.metric("Ingresos", f"{resultado['ingresos']:.2f} TN")
            col3.metric("Egresos", f"{resultado['egresos']:.2f} TN")
            col4.metric("Inv. Final", f"{resultado['inventario_final_calculado']:.2f} TN")
            
            if resultado["cuadra"]:
                st.success("✅ Inventario válido")
            else:
                st.error("❌ Inventario NO cuadra")

    if not validacion["cuadra"]:
        st.error("❌ Hay divisiones con inventario negativo. Revisa los datos.")
        st.stop()

    # Generar Batch SIREGAD
    if st.button("🚀 Generar Batch SIREGAD"):
        # Filtrar solo registros que van al batch (excluyendo inventarios inicial y final)
        df_batch = df_completo[
            (df_completo["IncludeBatch"] == True) & 
            (~df_completo["Concepto"].str.contains("inventario_inicial|inventario_final", na=False, regex=True))
        ].copy()
        
        if df_batch.empty:
            st.warning("No hay registros para incluir en el batch.")
            st.stop()
        
        # Asegurar que existan las columnas necesarias
        for col in ["Concentracion", "Expediente", "Rut", "Nombre"]:
            if col not in df_batch.columns:
                df_batch[col] = ""
        
        # Agrupar por "Grupo" sumando cantidades
        df_batch_agrupado = df_batch.groupby("Grupo", as_index=False).agg({
            "Cantidad": "sum",
            "Bodega": "first",
            "Material": "first",
            "Tipo Movimiento": "first",
            "Movimiento": "first",
            "Unidad": "first",
            "Concentracion": "first",
            "Expediente": "first",
            "Rut": "first",
            "Nombre": "first"
        }).copy()
        
        # Crear archivo Excel desde cero
        buffer_batch = io.BytesIO()
        
        with pd.ExcelWriter(buffer_batch, engine='openpyxl') as writer:
            df_dummy = pd.DataFrame()
            df_dummy.to_excel(writer, sheet_name='Datos', index=False)
            ws = writer.sheets['Datos']
            
            # Header amarillo
            columnas = ["dd-mm-aaaa", "Número Factura", "Id. Bodega", "SQC CAS/Nombre Mezcla", 
                       "Concentración (0,00000)", "Ingreso/Egreso", "Tipo de Movimiento", "Cantidad", 
                       "Unidad Medida", "Expediente Comunicación Anticipada", "Rut", "Nombre"]
            
            for col_idx, col_name in enumerate(columnas, 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.value = col_name
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", wrap_text=True)
            
            # Escribir datos sin colores
            row_num = 2
            for idx, row in df_batch_agrupado.iterrows():
                movimiento = row.get("Movimiento", "")
                tipo_mov = row.get("Tipo Movimiento", "")
                grupo = row.get("Grupo", "")
                
                # Generar número de factura
                if not (movimiento == "E" and tipo_mov == "VENT"):
                    if "ajuste" in grupo.lower():
                        numero_factura = f"'K.{movimiento}.Ajuste.DIC'"
                    else:
                        numero_factura = f"'K.{movimiento}.{tipo_mov}.DIC'"
                else:
                    numero_factura = f"'{grupo}'"
                
                ws.cell(row=row_num, column=1).value = fecha_ultimo_dia
                ws.cell(row=row_num, column=1).number_format = 'DD-MM-YYYY'
                ws.cell(row=row_num, column=2).value = numero_factura
                ws.cell(row=row_num, column=3).value = row.get("Bodega", "")
                ws.cell(row=row_num, column=4).value = row.get("Material", "")
                conc = row.get("Concentracion", "")
                ws.cell(row=row_num, column=5).value = conc if conc else "Ácido Sulfúrico (Conc: 96)"
                ws.cell(row=row_num, column=6).value = movimiento
                ws.cell(row=row_num, column=7).value = tipo_mov
                ws.cell(row=row_num, column=8).value = row.get("Cantidad", 0)
                ws.cell(row=row_num, column=8).number_format = '#,##0.000'
                ws.cell(row=row_num, column=9).value = row.get("Unidad", "TN")
                ws.cell(row=row_num, column=10).value = row.get("Expediente", "")
                rut = row.get("Rut", "")
                ws.cell(row=row_num, column=11).value = rut if rut else "61.704.000-K"
                nombre = row.get("Nombre", "")
                ws.cell(row=row_num, column=12).value = nombre if nombre else "Corporación Nacional del Cobre de Chile (Codelco)"
                
                row_num += 1
            
            # Ajustar anchos de columna
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 18
            ws.column_dimensions['E'].width = 25
            ws.column_dimensions['F'].width = 10
            ws.column_dimensions['G'].width = 18
            ws.column_dimensions['H'].width = 12
            ws.column_dimensions['I'].width = 12
            ws.column_dimensions['J'].width = 20
            ws.column_dimensions['K'].width = 15
            ws.column_dimensions['L'].width = 45

        st.success(f"✅ Batch generado con {len(df_batch_agrupado)} registros")
        
        st.download_button(
            "📥 Descargar Batch SIREGAD",
            buffer_batch,
            file_name=f"Batch_SIREGAD_{fecha_ultimo_dia.strftime('%Y-%m-%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

# =================== SECCIÓN DE VISUALIZACIÓN PERSISTENTE ===================
# Esta sección muestra los datos procesados incluso después de agregar ajustes manuales
# (cuando el botón "Cargar y Procesar" no fue presionado pero ya hay datos en session_state)

elif "df_completo" in st.session_state and st.session_state.df_completo is not None and not st.session_state.df_completo.empty:
    df_completo = st.session_state.df_completo
    
    st.subheader("📊 Datos extraídos (guardados)")
    
    if st.session_state.ajustes_manuales:
        st.info(f"📝 Incluye {len(st.session_state.ajustes_manuales)} ajuste(s) manual(es)")
    
    st.dataframe(df_completo, use_container_width=True)
    
    # Generar Excel intermedio con los datos actualizados
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from config import RUT_EMPRESAS
    
    buffer_intermedio = io.BytesIO()
    
    with pd.ExcelWriter(buffer_intermedio, engine='openpyxl') as writer:
        df_dummy = pd.DataFrame()
        df_dummy.to_excel(writer, sheet_name='Datos', index=False)
        ws = writer.sheets['Datos']
        
        row_actual = 1
        
        # Verde manzana pastel para todas las divisiones
        color_verde_manzana = "C1E1C1"
        colores_division = {
            "Salvador": color_verde_manzana, "Caletones": color_verde_manzana, "Potrerillos": color_verde_manzana,
            "TBA": color_verde_manzana, "San Antonio": color_verde_manzana, "Chuquicamata": color_verde_manzana,
            "Radomiro Tomic": color_verde_manzana, "Gabriela Mistral": color_verde_manzana,
            "Ministro Hales": color_verde_manzana, "Barquito": color_verde_manzana,
        }
        
        columnas = ["dd-mm-aaaa", "Número Factura", "Id. Bodega", "SQC CAS/Nombre Mezcla", 
                   "Concentración (0,00000)", "Ingreso/Egreso", "Tipo de Movimiento", "Cantidad", 
                   "Unidad Medida", "Expediente Comunicación Anticipada", "Rut", "Nombre"]
        
        # Orden de divisiones según especificación
        orden_divisiones = ["Potrerillos", "Caletones", "TBA", "Radomiro Tomic", 
                          "Chuquicamata", "Gabriela Mistral", "Ministro Hales", "Barquito", 
                          "San Antonio"]
        
        # Filtrar divisiones que existen en df_completo y mantener el orden
        divisiones_ordenadas = [div for div in orden_divisiones if div in df_completo["Division"].unique()]
        
        # ===== HEADER GLOBAL (UNA SOLA VEZ AL INICIO) =====
        for col_idx, col_name in enumerate(columnas, 1):
            cell = ws.cell(row=row_actual, column=col_idx)
            cell.value = col_name
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        row_actual += 1
        
        for division in divisiones_ordenadas:
            df_div = df_completo[df_completo["Division"] == division].copy()
            
            inv_inicial = df_div[df_div["Concepto"].str.contains("inventario_inicial", na=False)]["Cantidad"].sum()
            ingresos_total = df_div[df_div["Movimiento"] == "I"]["Cantidad"].sum()
            egresos_total = df_div[df_div["Movimiento"] == "E"]["Cantidad"].sum()
            inv_final_calculado = inv_inicial + ingresos_total - egresos_total
            inv_final_extraido = df_div[df_div["Concepto"].str.contains("inventario_final", na=False)]["Cantidad"].sum()
            
            color = colores_division.get(division, "FFFFFF")
            
            # Header de división
            ws[f'A{row_actual}'] = division.upper()
            ws.merge_cells(f'A{row_actual}:G{row_actual}')
            cell = ws[f'A{row_actual}']
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(horizontal="left")
            
            # INV INICIAL en columna H
            ws[f'H{row_actual}'] = inv_inicial
            ws[f'H{row_actual}'].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            ws[f'H{row_actual}'].font = Font(bold=True, size=11)
            ws[f'H{row_actual}'].alignment = Alignment(horizontal="center")
            ws[f'H{row_actual}'].number_format = '#,##0.000'
            row_actual += 1
            
            # Datos de movimientos - forzar conversión de IncludeBatch a booleano
            df_div["IncludeBatch"] = df_div["IncludeBatch"].astype(bool)
            df_batch = df_div[df_div["IncludeBatch"] == True].copy()
            
            for col in ["Concentracion", "Expediente", "Rut", "Nombre"]:
                if col not in df_batch.columns:
                    df_batch[col] = ""
            
            if not df_batch.empty:
                df_batch_agrupado = df_batch.groupby("Grupo", as_index=False).agg({
                    "Cantidad": "sum", "Bodega": "first", "Material": "first",
                    "Tipo Movimiento": "first", "Movimiento": "first", "Unidad": "first",
                    "Concentracion": "first", "Expediente": "first", "Rut": "first", "Nombre": "first"
                }).copy()
                
                for idx, row in df_batch_agrupado.iterrows():
                    ws.cell(row=row_actual, column=1).value = fecha_ultimo_dia
                    
                    # Columna 2: Número Factura - Generar código especial para NO VENT
                    movimiento = row.get("Movimiento", "")
                    tipo_mov = row.get("Tipo Movimiento", "")
                    grupo = row.get("Grupo", "")
                    
                    # Si NO es "E VENT", generar el código
                    if not (movimiento == "E" and tipo_mov == "VENT"):
                        # Si es ajuste (manual o inventario), usar 'Ajuste' en lugar del tipo_mov
                        if "ajuste" in grupo.lower():
                            numero_factura = f"'K.{movimiento}.Ajuste.DIC'"
                        else:
                            numero_factura = f"'K.{movimiento}.{tipo_mov}.DIC'"
                    else:
                        # Para E VENT, usar el grupo (número de factura real)
                        numero_factura = f"'{grupo}'"
                    
                    ws.cell(row=row_actual, column=2).value = numero_factura
                    ws.cell(row=row_actual, column=3).value = row.get("Bodega", "")
                    ws.cell(row=row_actual, column=4).value = row.get("Material", "")
                    conc = row.get("Concentracion", "")
                    ws.cell(row=row_actual, column=5).value = conc if conc else "Ácido Sulfúrico (Conc: 96)"
                    ws.cell(row=row_actual, column=6).value = row.get("Movimiento", "")
                    ws.cell(row=row_actual, column=7).value = row.get("Tipo Movimiento", "")
                    ws.cell(row=row_actual, column=8).value = row.get("Cantidad", 0)
                    ws.cell(row=row_actual, column=9).value = row.get("Unidad", "TN")
                    ws.cell(row=row_actual, column=10).value = row.get("Expediente", "")
                    rut = row.get("Rut", "")
                    ws.cell(row=row_actual, column=11).value = rut if rut else "61.704.000-K"
                    nombre = row.get("Nombre", "")
                    ws.cell(row=row_actual, column=12).value = nombre if nombre else "Corporación Nacional del Cobre de Chile (Codelco)"
                    
                    for col in range(1, 13):
                        cell = ws.cell(row=row_actual, column=col)
                        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                        if col == 8:
                            cell.number_format = '#,##0.000'
                        if col == 1:
                            cell.number_format = 'DD-MM-YYYY'
                    row_actual += 1
            
            # ===== TOTALES DE MOVIMIENTOS =====
            # Total de Ingresos (I)
            total_ingresos = df_div[(df_div["Movimiento"] == "I") & (~df_div["Concepto"].str.contains("inventario", na=False))]["Cantidad"].sum()
            
            cell_total_i_label = ws[f'A{row_actual}']
            cell_total_i_label.value = "TOTAL INGRESOS (I)"
            cell_total_i_label.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            cell_total_i_label.font = Font(bold=True)
            cell_total_i_label.alignment = Alignment(horizontal="right")
            ws.merge_cells(f'A{row_actual}:G{row_actual}')
            
            cell_total_i_valor = ws[f'H{row_actual}']
            cell_total_i_valor.value = total_ingresos
            cell_total_i_valor.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            cell_total_i_valor.font = Font(bold=True)
            cell_total_i_valor.alignment = Alignment(horizontal="center")
            cell_total_i_valor.number_format = '#,##0.000'
            
            row_actual += 1
            
            # Total de Egresos (E)
            total_egresos = df_div[(df_div["Movimiento"] == "E") & (~df_div["Concepto"].str.contains("inventario", na=False))]["Cantidad"].sum()
            
            cell_total_e_label = ws[f'A{row_actual}']
            cell_total_e_label.value = "TOTAL EGRESOS (E)"
            cell_total_e_label.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            cell_total_e_label.font = Font(bold=True)
            cell_total_e_label.alignment = Alignment(horizontal="right")
            ws.merge_cells(f'A{row_actual}:G{row_actual}')
            
            cell_total_e_valor = ws[f'H{row_actual}']
            cell_total_e_valor.value = total_egresos
            cell_total_e_valor.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            cell_total_e_valor.font = Font(bold=True)
            cell_total_e_valor.alignment = Alignment(horizontal="center")
            cell_total_e_valor.number_format = '#,##0.000'
            
            row_actual += 1
            
            # Total general (todos los movimientos sin inventarios)
            total_movimientos = df_div[~df_div["Concepto"].str.contains("inventario", na=False)]["Cantidad"].sum()
            
            cell_total_gen_label = ws[f'A{row_actual}']
            cell_total_gen_label.value = "TOTAL MOVIMIENTOS"
            cell_total_gen_label.fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
            cell_total_gen_label.font = Font(bold=True, color="FFFFFF")
            cell_total_gen_label.alignment = Alignment(horizontal="right")
            ws.merge_cells(f'A{row_actual}:G{row_actual}')
            
            cell_total_gen_valor = ws[f'H{row_actual}']
            cell_total_gen_valor.value = total_movimientos
            cell_total_gen_valor.fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
            cell_total_gen_valor.font = Font(bold=True, color="FFFFFF")
            cell_total_gen_valor.alignment = Alignment(horizontal="center")
            cell_total_gen_valor.number_format = '#,##0.000'
            
            row_actual += 1
            
            # Footer con inventarios finales
            ws[f'A{row_actual}'] = "INV FINAL (CALCULADO)"
            ws[f'A{row_actual}'].fill = PatternFill(start_color="FF69B4", end_color="FF69B4", fill_type="solid")
            ws[f'A{row_actual}'].font = Font(bold=True)
            ws.merge_cells(f'A{row_actual}:B{row_actual}')
            ws[f'C{row_actual}'] = inv_final_calculado
            ws[f'C{row_actual}'].fill = PatternFill(start_color="FF69B4", end_color="FF69B4", fill_type="solid")
            ws[f'C{row_actual}'].number_format = '#,##0.000'
            ws.merge_cells(f'C{row_actual}:D{row_actual}')
            
            ws[f'E{row_actual}'] = "INV FINAL (EXTRAÍDO)"
            ws[f'E{row_actual}'].fill = PatternFill(start_color="9370DB", end_color="9370DB", fill_type="solid")
            ws[f'E{row_actual}'].font = Font(bold=True)
            ws.merge_cells(f'E{row_actual}:F{row_actual}')
            ws[f'G{row_actual}'] = inv_final_extraido
            ws[f'G{row_actual}'].fill = PatternFill(start_color="9370DB", end_color="9370DB", fill_type="solid")
            ws[f'G{row_actual}'].number_format = '#,##0.000'
            ws.merge_cells(f'G{row_actual}:H{row_actual}')
            row_actual += 2
        
        # ===== SECCIÓN TERMINALES MEJILLONES =====
        color_terminales = "FFD700"  # Dorado para terminales
        
        # Header de Terminales Mejillones
        ws[f'A{row_actual}'] = "TERMINALES MEJILLONES"
        ws.merge_cells(f'A{row_actual}:L{row_actual}')
        cell = ws[f'A{row_actual}']
        cell.fill = PatternFill(start_color=color_terminales, end_color=color_terminales, fill_type="solid")
        cell.font = Font(bold=True, size=14)
        cell.alignment = Alignment(horizontal="center")
        row_actual += 1
        
        # Encabezados de columnas
        for col_idx, col_name in enumerate(columnas, 1):
            cell = ws.cell(row=row_actual, column=col_idx)
            cell.value = col_name
            cell.fill = PatternFill(start_color=color_terminales, end_color=color_terminales, fill_type="solid")
            cell.font = Font(bold=True, color="000000")
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        row_actual += 1
        
        # Datos de los 3 terminales
        terminales_mejillones = [
            {"nombre": "Interacid", "bodega": "MEJICL", "cantidad": 64249.702},
            {"nombre": "TPM", "bodega": "MEJ", "cantidad": 8772.753},
            {"nombre": "Terquim", "bodega": "TERMEJ", "cantidad": 5598.250},
        ]
        
        for terminal in terminales_mejillones:
            ws.cell(row=row_actual, column=1).value = fecha_ultimo_dia
            ws.cell(row=row_actual, column=2).value = f"K.E.TIEB.DIC"
            ws.cell(row=row_actual, column=3).value = terminal["bodega"]
            ws.cell(row=row_actual, column=4).value = "7664-93-9"
            ws.cell(row=row_actual, column=5).value = "Ácido Sulfúrico (Conc: 96)"
            ws.cell(row=row_actual, column=6).value = "E"
            ws.cell(row=row_actual, column=7).value = "TIEB"
            ws.cell(row=row_actual, column=8).value = terminal["cantidad"]
            ws.cell(row=row_actual, column=9).value = "TN"
            ws.cell(row=row_actual, column=10).value = ""
            ws.cell(row=row_actual, column=11).value = "61.704.000-K"
            ws.cell(row=row_actual, column=12).value = "Corporación Nacional del Cobre de Chile (Codelco)"
            
            for col in range(1, 13):
                cell = ws.cell(row=row_actual, column=col)
                cell.fill = PatternFill(start_color=color_terminales, end_color=color_terminales, fill_type="solid")
                if col == 8:
                    cell.number_format = '#,##0.000'
                if col == 1:
                    cell.number_format = 'DD-MM-YYYY'
            
            row_actual += 1
        
        row_actual += 2
        
        # Ajustar ancho de columnas
        for col_idx in range(1, 13):
            ws.column_dimensions[get_column_letter(col_idx)].width = 15
    
    buffer_intermedio.seek(0)
    
    st.download_button(
        "📥 Descargar Excel Intermedio (Actualizado)",
        buffer_intermedio,
        file_name=f"Intermedio_SIREGAD_{fecha_ultimo_dia.strftime('%Y-%m-%d')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="download_intermedio_persistente"
    )
    
    # Validación
    validacion = validar_inventario(df_completo)
    
    st.subheader("🔎 Validación por División")
    for resultado in validacion["por_division"]:
        with st.expander(f"División: {resultado['division']}", expanded=True):
            col1, col2, col3, col4 = st.columns(4)
            col1.metric("Inv. Inicial", f"{resultado['inventario_inicial']:.2f} TN")
            col2.metric("Ingresos", f"{resultado['ingresos']:.2f} TN")
            col3.metric("Egresos", f"{resultado['egresos']:.2f} TN")
            col4.metric("Inv. Final", f"{resultado['inventario_final_calculado']:.2f} TN")
            
            if resultado["cuadra"]:
                st.success("✅ Inventario válido")
            else:
                st.error("❌ Inventario NO cuadra")
    
    if not validacion["cuadra"]:
        st.warning("⚠️ Hay divisiones con inventario negativo. Revisa los datos.")
    
    # Botón para generar Batch
    if st.button("🚀 Generar Batch SIREGAD", key="btn_batch_persistente"):
        # Filtrar solo registros que van al batch (excluyendo inventarios)
        df_batch = df_completo[
            (df_completo["IncludeBatch"] == True) & 
            (~df_completo["Concepto"].str.contains("inventario_inicial|inventario_final", na=False, regex=True))
        ].copy()
        
        if df_batch.empty:
            st.warning("No hay registros para incluir en el batch.")
        else:
            # Asegurar que existan las columnas necesarias
            for col in ["Concentracion", "Expediente", "Rut", "Nombre"]:
                if col not in df_batch.columns:
                    df_batch[col] = ""
            
            # Agrupar por "Grupo" sumando cantidades
            df_batch_agrupado = df_batch.groupby("Grupo", as_index=False).agg({
                "Cantidad": "sum",
                "Bodega": "first",
                "Material": "first",
                "Tipo Movimiento": "first",
                "Movimiento": "first",
                "Unidad": "first",
                "Concentracion": "first",
                "Expediente": "first",
                "Rut": "first",
                "Nombre": "first"
            }).copy()
            
            # Crear archivo Excel desde cero
            buffer_batch = io.BytesIO()
            
            with pd.ExcelWriter(buffer_batch, engine='openpyxl') as writer:
                df_dummy = pd.DataFrame()
                df_dummy.to_excel(writer, sheet_name='Datos', index=False)
                ws = writer.sheets['Datos']
                
                # Header amarillo
                columnas = ["dd-mm-aaaa", "Número Factura", "Id. Bodega", "SQC CAS/Nombre Mezcla", 
                           "Concentración (0,00000)", "Ingreso/Egreso", "Tipo de Movimiento", "Cantidad", 
                           "Unidad Medida", "Expediente Comunicación Anticipada", "Rut", "Nombre"]
                
                for col_idx, col_name in enumerate(columnas, 1):
                    cell = ws.cell(row=1, column=col_idx)
                    cell.value = col_name
                    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="center", wrap_text=True)
                
                # Escribir datos sin colores
                row_num = 2
                for idx, row in df_batch_agrupado.iterrows():
                    movimiento = row.get("Movimiento", "")
                    tipo_mov = row.get("Tipo Movimiento", "")
                    grupo = row.get("Grupo", "")
                    
                    # Generar número de factura
                    if not (movimiento == "E" and tipo_mov == "VENT"):
                        if "ajuste" in grupo.lower():
                            numero_factura = f"'K.{movimiento}.Ajuste.DIC'"
                        else:
                            numero_factura = f"'K.{movimiento}.{tipo_mov}.DIC'"
                    else:
                        numero_factura = f"'{grupo}'"
                    
                    ws.cell(row=row_num, column=1).value = fecha_ultimo_dia
                    ws.cell(row=row_num, column=1).number_format = 'DD-MM-YYYY'
                    ws.cell(row=row_num, column=2).value = numero_factura
                    ws.cell(row=row_num, column=3).value = row.get("Bodega", "")
                    ws.cell(row=row_num, column=4).value = row.get("Material", "")
                    conc = row.get("Concentracion", "")
                    ws.cell(row=row_num, column=5).value = conc if conc else "Ácido Sulfúrico (Conc: 96)"
                    ws.cell(row=row_num, column=6).value = movimiento
                    ws.cell(row=row_num, column=7).value = tipo_mov
                    ws.cell(row=row_num, column=8).value = row.get("Cantidad", 0)
                    ws.cell(row=row_num, column=8).number_format = '#,##0.000'
                    ws.cell(row=row_num, column=9).value = row.get("Unidad", "TN")
                    ws.cell(row=row_num, column=10).value = row.get("Expediente", "")
                    rut = row.get("Rut", "")
                    ws.cell(row=row_num, column=11).value = rut if rut else "61.704.000-K"
                    nombre = row.get("Nombre", "")
                    ws.cell(row=row_num, column=12).value = nombre if nombre else "Corporación Nacional del Cobre de Chile (Codelco)"
                    
                    row_num += 1
                
                # Ajustar anchos de columna
                ws.column_dimensions['A'].width = 12
                ws.column_dimensions['B'].width = 20
                ws.column_dimensions['C'].width = 12
                ws.column_dimensions['D'].width = 18
                ws.column_dimensions['E'].width = 25
                ws.column_dimensions['F'].width = 10
                ws.column_dimensions['G'].width = 18
                ws.column_dimensions['H'].width = 12
                ws.column_dimensions['I'].width = 12
                ws.column_dimensions['J'].width = 20
                ws.column_dimensions['K'].width = 15
                ws.column_dimensions['L'].width = 45
            
            st.success(f"✅ Batch generado con {len(df_batch_agrupado)} registros")
            
            st.download_button(
                "📥 Descargar Batch SIREGAD",
                buffer_batch,
                file_name=f"Batch_SIREGAD_{fecha_ultimo_dia.strftime('%Y-%m-%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="download_batch_persistente"
            )
